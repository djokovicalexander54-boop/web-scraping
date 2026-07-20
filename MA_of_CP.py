i=2
word_list = []
word = str(input('please enter data 1:'))
word_list.append(word)
while word!= "":
    word = str(input(f'please enter data {i}:'))
    if word != "":
        word_list.append(word)
        i+=1
print(word_list)
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import telebot
from telebot import apihelper
from bs4 import BeautifulSoup
import time
import os
import sqlite3
from openai import OpenAI
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
hj = sqlite3.connect(r"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices\data_table.db")
conn = hj.cursor()
conn.execute('''
        CREATE TABLE IF NOT EXISTS products (
        product_name TEXT PRIMARY KEY, 
        price TEXT,
        exist TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
''')
hj.commit()
speed = float(input('attentio!!! please choose the speed of data extraction? \n slow mode --> (5s) \n normal and safe mode --> (3s) \n fast mode --> (1s) \n very fast mode! --> (0.05s) \n for example, if you wanna fast, please write just number 1:'))
time.sleep(1)
print('OK, please wait..')
Service = Service(ChromeDriverManager().install())
site = webdriver.Chrome(service=Service)
L_dir = r"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices"  
gg = os.path.join(L_dir, 'excel_files')
if not os.path.exists(gg):
    os.makedirs(gg)
name_list = []
price_list = []
exist_list = []
normal_list = []
i=0
zxxx_list = []
excel_list = []
for word in word_list:
    name_list1 = []
    price_list1 = []
    exist_list1 = []
    site.get("https://www.digikala.com/")
    element1 = WebDriverWait(site, 100).until( 
    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text'][name='search-input']"))
    )
    element1.click()
    action = ActionChains(site)
    action.move_to_element(element1)
    action.click()
    action.send_keys(word)
    action.send_keys(Keys.RETURN)
    action.perform()
    WebDriverWait(site, 100).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='product-card']"))
    )
    k=1
    while k==1:
        gg = site.execute_script("return window.pageYOffset;")
        site.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        kk = site.execute_script("return window.pageYOffset;")
        if kk==gg:
            number = len(site.find_elements(By.CSS_SELECTOR, "div[data-testid='product-card']"))
            site.execute_script("window.scrollBy(0,-3000)")
            WebDriverWait(site, 100).until( 
            lambda site: len(site.find_elements(By.CSS_SELECTOR, "div[data-testid='product-card']")) >= number
            )
            time.sleep(1)
            site.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            ll = site.execute_script("return window.pageYOffset;")
            if ll==kk:
                k=2
    site_html = BeautifulSoup(site.page_source, 'html.parser')
    all_item = site_html.find_all('div', class_="flex grow relative flex-col")
    zxx_list = []
    r=1
    for item in all_item:
        name = item.select_one('h3[class*="ellipsis-2 text-body2-strong text-neutral-700 styles"]')
        if name:
            name_list.append(name.text.strip())
            name_list1.append(name.text.strip())
            n = name.text.strip()
            if r<=3:
                site.execute_script("window.scrollBy(0,-44000)")
                time.sleep(speed)
                card = site.find_elements(By.CSS_SELECTOR, "div[data-testid='product-card']")
                number = card[r-1]
                number.click()
                time.sleep(1)
                q=1
                while q==1:
                    gg = site.execute_script("return window.pageYOffset;")
                    site.execute_script("window.scrollBy(0,100)")
                    kk = site.execute_script("return window.pageYOffset;")
                    time.sleep(speed)
                    if gg==kk:
                        q=2
                site_html00 = BeautifulSoup(site.page_source, 'html.parser')
                ff_list = site_html00.find_all('p', class_="text-body-1 text-neutral-900 mb-1 break-words")
                zx_list = []
                for comment in ff_list:
                    zx_list.append(comment.text.strip())
                zxx_list.append(zx_list)
                site.back()
                time.sleep(3)
            r+=1
        else:
            n = 'there is not name!!'
            name_list.append('there is not name!!')
            name_list1.append('there is not name!!')
        price = item.find('span', attrs={'data-testid':'price-final'})
        if price:
            p = price.text.strip()
            price_list.append(f"{price.text.strip()} تومان")
            price_list1.append(f"{price.text.strip()} تومان")
        else:
            price_list.append('there is not price!!')
            price_list1.append('there is not price!!')
            p = 'there is not price!!'
        exist = item.find('p', class_="text-caption text-primary-700")
        if exist:
            exist_list.append(exist.text.strip())
            exist_list1.append(exist.text.strip())
            e = exist.text.strip()
        else:
            exist_list.append('موجودی به اندازه کافی هست')
            exist_list1.append('موجودی به اندازه کافی هست')
            e = 'موجودی به اندازه کافی هست'
        conn.execute(
            "INSERT OR REPLACE INTO products (product_name, price, exist) VALUES (?,?,?)",
            (n,p,e)
        )
        hj.commit()
    name_list.append(f"end of product")
    price_list.append('---')
    exist_list.append('---')
    u=0
    list_df = []
    while u<len(name_list1):
        list_df.append(f"{name_list1[u]} ---> {price_list1[u]} : {exist_list1[u]}")
        u+=1
    normal_list.append(list_df)
    zxxx_list.append(zxx_list)
    small_dict = {f"name {word}":name_list1, 'price': price_list1, 'exist': exist_list1}
    small_table = pd.DataFrame(small_dict)
    document = fr"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices\excel_files\excel {word}.xlsx"
    small_table.to_excel(document)
    excel_list.append(document)
    print(f'scrapting {word}')
conn.execute("SELECT * FROM products")
all_rows = conn.fetchall()
conn.execute("SELECT product_name, price, exist FROM products WHERE product_name LIKE ?", (f"%{'آیفون'}%",))
sdf = conn.fetchall()
hj.commit()
hj.close()
site.quit()
my_dict = {'name product':name_list, 'price product': price_list, 'exist': exist_list}
table = pd.DataFrame(my_dict)
table.to_excel(r"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices\table_products.xlsx", index=False)
file = openpyxl.load_workbook(r"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices\table_products.xlsx")
sheet = file.active
font_style = Font(size=24)
font_style1 = Font(size=18)
fill_style = PatternFill(start_color="FFFF00" , end_color="FFFF00", fill_type="solid")
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == 'name product':
            cell.font = font_style
        elif cell.value == 'price product':
            cell.font = font_style
        else:
            if cell.value == 'exist':
                cell.font = font_style
            elif cell.value == "end of product":
                cell.fill = fill_style
                cell.font = font_style1
            else:
                if cell.value == '---':
                    cell.fill = fill_style
                    cell.font = font_style1
file.save(r"C:\Users\karino\Desktop\PROJECTS\monitoring_&_analyzing_of_prices\table_products.xlsx")
time.sleep(10)
TOKON = "توکن موردنظر جهت ساخت ربات تلگرام"
proxy = 'پروکسی موردنظر جهت رفع فیلترینگ تلگرام'
apihelper.proxy = {'https': proxy, 'http': proxy}
bot = telebot.TeleBot(TOKON)    
async def data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard =[]
    for ds in word_list:
        kl = word_list.index(ds)
        button = [InlineKeyboardButton(text=f"نمونه {ds}", callback_data=f"show_{kl}"),InlineKeyboardButton(text=f"فایل اکسل همه {ds}", callback_data=f"see_{kl}")]
        keyboard.append(button)
    reply = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("which one data you wanna?", reply_markup=reply)
async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mm = update.callback_query
    await mm.answer()
    data = mm.data
    if data.startswith("show_"):
        cb = int(data.split("_")[1])
        select = normal_list[cb] 
        text_reply = "information that you want"
        j=1
        analyze_list = []
        ux_list = zxxx_list[cb]
        for item in select:
            if j!=3:
                text_reply=f"{item}"
                analyze_list.append(text_reply)
                await context.bot.send_message(chat_id=update.effective_chat.id, text=text_reply)
                cv = OpenAI(
                    base_url = "آدرس هوش مصنوعی موردنظر",
                    api_key = "رمز عبور هوش مصنوعی موردنظر"
                )
                sentence = f"please analys this coments list for one product {ux_list[j-1]}.and say that if i buy this product is good idea or no. explain that analys in 3 or 4 line. just tell me to persion language.."
                responce = cv.chat.completions.create( 
                    model = "مدل هوش مصنوعی موردنظر", 
                    messages=[{"role":"user", "content":sentence}] 
                )
                ai_answer = responce.choices[0].message.content
                await context.bot.send_message(chat_id=update.effective_chat.id, text=ai_answer)
                await context.bot.send_message(chat_id=update.effective_chat.id, text='********************************')
                j+=1
        cv = OpenAI(
            base_url = "آدرس هوش مصنوعی موردنظر",
            api_key = "رمز عبور هوش مصنوعی موردنظر"
        )
        sentence = f"please analys this prices {analyze_list}, so explain that analys in 3 or 4 line. also please suggest that now buy one of that is good idea or no! tell me to persion language.. if there is no comment in this list, please just write 'هیچ کامنتی برای این محصول موجود نیست' and so do not write every things."
        responce = cv.chat.completions.create( 
            model = "مدل هوش مصنوعی موردنظر", 
            messages=[{"role":"user", "content":sentence}] 
        )
        ai_answer = responce.choices[0].message.content
        await context.bot.send_message(chat_id=update.effective_chat.id, text="پیشنهاد کلی برای کیفیت این دسته محصولات دیجی کالا با توجه به نظرات:")
        await context.bot.send_message(chat_id=update.effective_chat.id, text=ai_answer)
        await context.bot.send_message(chat_id=update.effective_chat.id, text='-------------------------------------------------------')
    elif data.startswith("see_"):
        cb = int(data.split("_")[1])
        ty = excel_list[cb]
        await context.bot.send_document(chat_id=update.effective_chat.id, document=ty, caption="این فایل اکسل، مشخصات محصول موردنظر را نمایش میدهد")
if __name__ == '__main__':
    app = Application.builder().token(TOKON).build()
    app.add_handler(CommandHandler("start", data))
    app.add_handler(CallbackQueryHandler(button))
    print('yes')
    app.run_polling()                       